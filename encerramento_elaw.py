from tkinter import INSERT
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import time
from time import sleep
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options  
from selenium.webdriver.support.wait import WebDriverWait
import os
import openpyxl
import shutil
import threading
from printlog import print_log as prt
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import os, openpyxl, shutil, threading, pathlib, subprocess, sys


tipos_encerramento = {
    "resultado": {
        "Acordo": "Acordo",
        "Procedente": "Procedente",
        "Parcialmente Procedente": "Procedente em Parte",
        "Improcedente": "Improcedente",
        "Extinto sem resolução do mérito": "Extinto o Processo sem Apreciação do Mérito",
        "Extinto por desistência da parte": "Desistência da Ação",
        "Extinto por ausência da parte": "Extinto por Ausência do Autor"
    },
    "motivo": {
        "Acordo": "Acordo (Acordo)",
        "Procedente": "Julgada Procedente (Procedente)",
        "Parcialmente Procedente": "Julgada Procedente (Procedente em Parte)",
        "Improcedente": "Julgada Improcedente (Improcedente)",
        "Extinto sem resolução do mérito": "Extinto o Processo sem Apreciação do Mérito",
        "Extinto por desistência da parte": "Extinto (Desistência Da Ação)",
        "Extinto por ausência da parte": "Extinto por Ausência do Autor"
    }
}

class ELAWCrawler:

    def __init__(self, input_file , credentials, tipo_encerramento = 'Acordo'):
        
        chrome_options = Options()
        #chrome_options.add_argument("--headless")
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
        wait = WebDriverWait(driver, 10)

        chrome_options.add_argument("--window-size=1920,1080")
        login, password = credentials.split('>!>')
        self.credentials = {
            'login': login,
            'password': password,
        }
        self.input_file = input_file
        self.output_dir_path = pathlib.Path(input_file).parent.resolve()
        self.current_dir = pathlib.Path(__file__).parent.resolve()
        self.tipo_encerramento = tipo_encerramento
        self.resultado = tipos_encerramento['resultado'].get(tipo_encerramento)
        self.motivo = tipos_encerramento['motivo'].get(tipo_encerramento)
        wait = WebDriverWait(driver, 10)
        
        try:
            get_login_elaw = threading.Thread(target=self.login_elaw, args=(driver, wait))
            get_login_elaw.start()
            #self.get_login_and_password(driver)
        except:
            prt(type='error',message='Erro ao fazer login')

    def login_elaw(self, driver, login_data, wait):

        driver.get("https://amazonas.elaw.com.br/login.elaw")

        prt(type='log', message='Fazendo login no sistema...')

        # wait until page load
        username = wait.until(EC.presence_of_element_located((By.ID,'username')))
        username.click()
        login = login_data['username']
        username.send_keys(login)


        password = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, '#password')))
        password.click()
        senha = login_data['password']
        password.send_keys(senha)

        entrar = wait.until(EC.presence_of_element_located((By.ID, "j_id_a_1_5_f")))
        entrar.click()
        prt(type='log', message='Login efetuado com sucesso')

        try:
            self.get_processos_from_input(driver=driver, wait=wait)
        except:
            prt(type='error', message='Erro ao acessar processos do input')
            sleep(1)
       
    def search_processo(self, driver, processo_data, row, wait):

        driver.get("https://amazonas.elaw.com.br/processoList.elaw")

        prt(type='log',message='Buscando processo {}...'.format(processo_data['n_processo']), row=row)

        numproc = wait.until(EC.presence_of_element_located((By.ID,'tabSearchTab:txtSearch')))
        numproc.click()
        numproc.clear()
        numproc.send_keys(processo_data['n_processo'])

        search_button = driver.find_element(By.ID, 'btnPesquisar')
        search_button.click()

        search_result = wait.until(EC.presence_of_element_located((By.ID,'dtProcessoResults_data')))
        open_proc = search_result.find_element(By.ID, 'dtProcessoResults:0:btnProcesso')
        
        if open_proc is not None:           
            open_proc.click()
            prt(type='log', message='Processo encontrado', row=row)
            self.encerrar_processo(driver, processo_data, row=row, wait=wait)
        
        else:
            prt(type='error', message='Erro ao preencher dados do encerramento do processo {}'.format(processo_data['n_processo']), row=row)
            self.append_error_on_output(n_processo=processo_data['n_processo'], error='Erro ao preencher dados do encerramento')

    def encerrar_processo(self, driver, processo_data, wait, row):

        try:
            encerrar_button = wait.until(EC.element_to_be_clickable((By.ID,'btnEncerrarProcesso')))
            encerrar_button.click()
        except Exception as e:
            prt(type='error', message='Erro ao encerrar processo, talvez ele já esteja encerrado', row=row)
            self.append_error_on_output(n_processo=processo_data['n_processo'], error='rro ao encerrar processo, talvez ele já esteja encerrado')
            return

        prt(type='log', message='Preenchendo dados do encerramento...', row=row)
        sleep(2)
        
        resultado_button = wait.until(EC.presence_of_element_located((By.ID,'j_id_2p')))
        resultado_button.click()
        sleep(3)

        resultado_option = wait.until(EC.presence_of_element_located((By.XPATH,'//ul[@id="j_id_2p_items"]//li[@data-label="{}"]'.format(self.resultado))))
        resultado_option.click()
        sleep(3)

        motivo_encerramento_button = wait.until(EC.presence_of_element_located((By.ID,'j_id_2w')))
        motivo_encerramento_button.click()
        sleep(3)

        motivo_encerramento_option = wait.until(EC.presence_of_element_located((By.XPATH,'//ul[@id="j_id_2w_items"]//li[@data-label="{}"]'.format(self.motivo))))
        motivo_encerramento_option.click()
        sleep(3)

        valor = wait.until(EC.presence_of_element_located((By.XPATH,'//*[@id="j_id_7m_5_1_1:j_id_7m_5_1_1f_input"]')))
        valor.click()
        sleep(3)
        valor.send_keys('{value}{dot}'.format(dot='.',value=processo_data['valor']).split('.')[0].replace(',',''))

        driver.execute_script("document.querySelector('input.ui-inputfield.ui-inputtext.ui-widget.ui-state-default.ui-corner-all').blur()")

        prt(type='log', message='Dados preenchidos com sucesso', row=row)
        sleep(3)

        encerrar_form_button = wait.until(EC.presence_of_element_located((By.ID,'btnEncerrarProcessoDialog')))
        encerrar_form_button.click()

        prt(type='log', message='Processo {} encerrado'.format(processo_data['n_processo']), row=row)
        sleep(1)

    def get_processos_from_input(self, driver, wait):

        prt(type='log', message='Carregando processos do input..')
        # copy errors spreadsheet
        try: 
            example_path = os.path.join(self.current_dir, 'encerramentos_elaw_erros.xlsx')
            example_copy_path = '{dir_path}/encerramentos_elaw_erros.xlsx'.format(dir_path=self.output_dir_path)
            shutil.copyfile(example_path, example_copy_path)

        except:
            prt(type=='error',message='Erro ao criar planilha com os processos que não puderem ser encerrados')

        # iterate from each process on input
        #input_filename = os.path.join(os.getcwd(),'inputsearch', 'processos_elaw.xlsx')
        input_filename = self.input_file
        wrkbk_input = openpyxl.load_workbook(filename=input_filename)
        sheet_input = wrkbk_input.active
        prt(type='log',message='Processos carregados')

        for i in range(2, sheet_input.max_row+1):
            cell_obj = sheet_input.cell(row=i, column=1)
            if cell_obj.value is not None and cell_obj.value != '':
                processo_data = {}
                processo_data['n_processo'] = cell_obj.value.replace(' ','')
                processo_data['sentença'] = sheet_input.cell(row=i, column=2).value
                processo_data['valor'] = str(sheet_input.cell(row=i, column=3).value)

                try:
                    self.search_processo(driver=driver, processo_data=processo_data, row=i-1, wait=wait)
                except:
                    prt(type='error',message='Erro ao buscar processo {}'.format(processo_data['n_processo']))
                    self.append_error_on_output(n_processo=processo_data['n_processo'], error='Erro ao buscar processo')

            if i == sheet_input.max_row:
                    prt(type='log', message='Fim dos processos')
                    try:
                        if sys.platform == 'linux':
                            os.system('xdg-open "%s"' % self.output_dir_path)
                            subprocess.call(["xdg-open", '{dir_path}/encerramentos_elaw_erros.xlsx'.format(dir_path=self.output_dir_path)])
                        else:
                            os.system(f'start {self.output_dir_path}')
                            os.startfile('{dir_path}/encerramentos_elaw_erros.xlsx'.format(dir_path=self.output_dir_path))
                    except Exception as e:
                        prt(type='error', message=e)

    def append_error_on_output(self, n_processo, error):
        try:
            errors_copy_path = '{dir_path}/encerramentos_elaw_erros.xlsx'.format(dir_path=self.output_dir_path)
            output_filename = errors_copy_path

            wb = openpyxl.load_workbook(filename=output_filename)
            sheet = wb['Sheet1']
            append_data = [n_processo, error]

            sheet.append(append_data)
            wb.save(output_filename)
        except:
            prt(type='error',message='Erro ao salvar o "tipo de erro" do processo')


ELAWCrawler(
    credentials=sys.argv[1],
    input_file=sys.argv[2],
    tipo_encerramento=sys.argv[3]
)
