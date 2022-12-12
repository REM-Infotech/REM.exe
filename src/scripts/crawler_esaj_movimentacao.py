from selenium import webdriver
from time import sleep
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options  
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import os
import json
import openpyxl
import shutil
from termcolor import colored
import threading
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from printlog import print_log as prt
import pyautogui
from selenium.webdriver.support.ui import Select
import sys
import pathlib

class ESAJCrawler:

    def __init__(self, credentials, input_file, output_dir_path):

        chrome_options = Options()
        path_websigner = os.path.join(pathlib.Path(__file__).parent.resolve(), 'Web-Signer.crx')
        chrome_options.add_extension(path_websigner)
        #chrome_options.add_argument("--headless")
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
        wait = WebDriverWait(driver, 10)

        login, password = credentials.split('>!>')
        self.credentials = {
            'login': login,
            'password': password
        }
        self.input_file = input_file
        self.output_dir_path = pathlib.Path(input_file).parent.resolve()
        self.current_dir = pathlib.Path(__file__).parent.resolve()
        self.output_planilha_copia = None
 
        try:
            login_esaj = threading.Thread(target=self.login_esaj, args=(driver, wait))
            login_esaj.start()
            #self.get_login_and_password(driver)
        except:
            prt(type="error", message='Erro ao recuperar dados de login e senha', row=1)

    def login_esaj(self, driver, wait):
        # 0638164-88.2019.8.04.0015
        # driver.get("https://consultasaj.tjam.jus.br/sajcas/login#aba-certificado")
        driver.get("https://consultasaj.tjam.jus.br/sajcas/login#aba-cpf")
        sleep(3)
        login = self.credentials['login']
        
        logincert = wait.until(EC.presence_of_all_elements_located((By.XPATH, '//*[@id="certificados"]')))
        for cert in logincert:
            loginopt = cert.find_elements(By.TAG_NAME, "option")
            sleep(5)
            for option in loginopt:
                    if login in option.text.lower():
                        try:
                            sencert = option.get_attribute("value")
                            select = Select(driver.find_element(By.CSS_SELECTOR, 'select[id="certificados"]'))
                            select.select_by_value(sencert)
                            entrar = driver.find_element(By.XPATH, '//*[@id="submitCertificado"]')
                            entrar.click()
                            pyautogui.click(768,573,duration=2)
                            pyautogui.click(1002,638,duration=2)
                            self.websigner_confirmed = 1
                            prt(type='log', message='Login efetuado com sucesso', row=1)
                            sleep(2)
                        except Exception as e:
                            print(e)

                        try:
                            self.get_input_process(driver, wait=wait)
                        except:
                            print(colored('Erro ao recuperar processos do input', 'red'))
                            prt(type='error', message='Erro ao recuperar processos do input', row=1)

    def get_processo(self, driver, processo_data, row, wait):

        n_processo = processo_data[0]

        driver.get("https://consultasaj.tjam.jus.br/cpopg/open.do")

        npu = driver.find_element(By.ID, 'radioNumeroAntigo')
        npu.click()

        input_proc_1grau = driver.find_element(By.ID, 'nuProcessoAntigoFormatado')

        input_proc_1grau.send_keys(n_processo)

        consultar = driver.find_element(By.ID,'botaoConsultarProcessos')
        consultar.click()
        # sleep(60)

        try:
            self.get_process_informations(driver=driver, processo_data=processo_data, row=row, wait=wait)
        except:
            print(colored('Erro ao buscar informações do processo {}'.format(n_processo), 'red'))
            prt(type="error", message='Erro ao buscar informações do processo {}'.format(n_processo), row=row)
        
    def get_process_informations(self, driver, processo_data, row, wait):

        #sleep(10)

        n_processo = processo_data[0]

        more_details = driver.find_element(By.XPATH, '//a[@href="#maisDetalhes"]')
        more_details.click()
        sleep(1)

        competencia = driver.find_element(By.ID, 'foroProcesso').text
        vara = driver.find_element(By.ID, 'varaProcesso').text
        dist = driver.find_element(By.ID, 'dataHoraDistribuicaoProcesso').text

        table_partes = driver.find_element(By.ID, 'tablePartesPrincipais')

        polo_ativo = table_partes.find_elements(By.TAG_NAME, 'tr')[0].find_elements(By.TAG_NAME, 'td')[1].text.split('\n')[0]
        cpf_polo_ativo = 'Não consta'
        try:
            adv_polo_ativo = table_partes.find_elements(By.TAG_NAME, 'tr')[0].find_elements(By.TAG_NAME, 'td')[1].text.split(':')[1].replace('Advogada','').replace('Advogado','')
        except:
            adv_polo_ativo = 'Não consta'

        polo_passivo = table_partes.find_elements(By.TAG_NAME, 'tr')[1].find_elements(By.TAG_NAME, 'td')[1].text.split('\n')[0]
        cpf_polo_passivo = 'Não consta'

        try:
            adv_polo_passivo = table_partes.find_elements(By.TAG_NAME, 'tr')[1].find_elements(By.TAG_NAME, 'td')[1].text.split(':')[1].replace('Advogada','').replace('Advogado','')
        except:
            adv_polo_passivo = 'Não consta'

        valor = driver.find_element(By.ID, 'valorAcaoProcesso').text

        processo_data = [n_processo, competencia, vara, dist, polo_ativo, cpf_polo_ativo,adv_polo_ativo, polo_passivo, cpf_polo_passivo,adv_polo_passivo, valor, processo_data[1]]

        try:
            self.get_moves(driver=driver, process_data=processo_data, row=row, wait=wait)
        except:
            print(colored('Não há movimentação para o processo {}'.format(n_processo), 'red'))
            prt(type="error", message='Não há movimentação para o processo {}'.format(n_processo), row=row)
            try:
                self.append_process_on_output(processo_data=processo_data)
            except:
                print(colored('Erro ao adicionar processo {} na planilha'.format(processo_data[0]), 'red'))
                prt(type="error", message='Erro ao adicionar processo {} na planilha'.format(processo_data[0]), row=row)

    def get_moves(self, driver, process_data, row, wait):
        
        print(colored('Salvando movimentações...', 'yellow'))
        prt(type="log", message='Salvando movimentações...', row=row)

        #0638164-88.2019.8.04.0015
        sleep(3)
        try: 
            table_all_moves = driver.find_element(By.ID, 'tabelaTodasMovimentacoes')
            driver.execute_script('document.querySelector("#tabelaTodasMovimentacoes").style.display = "block"')
        except:
            table_all_moves = driver.find_element(By.ID, 'tabelaUltimasMovimentacoes')
            driver.execute_script('document.querySelector("#tabelaUltimasMovimentacoes").style.display = "block"')

        all_moves = []
        row_move = table_all_moves.find_elements(By.CSS_SELECTOR, '.containerMovimentacao')

        for move in row_move:
            move_date = move.find_element(By.CLASS_NAME, 'dataMovimentacao').text
            move_description = move.find_element(By.CLASS_NAME, 'descricaoMovimentacao').text

            for termo in process_data[11]:
                if termo.lower() in move_description.lower():
                    all_moves.append('{move_date} - {move_description}'.format(move_date=move_date,move_description=move_description))

        all_moves_string = '\n'.join(all_moves)
        process_data.append(','.join(process_data[11]))
        process_data[11] = all_moves_string

        try:
            self.append_process_on_output(processo_data=process_data, row=row)
        except:
            print(colored('Erro ao adicionar processo {} na planilha'.format(process_data[0]), 'red'))
            prt(type="error", message='Erro ao adicionar processo {} na planilha'.format(process_data[0]), row=row)

    # def get_login_and_password(self, driver, wait):

    #     login_file = os.path.join(os.getcwd(),'inputsearch', 'login_dados.json')

    #     f = open(login_file)
    #     login_data = json.load(f)
    #     login_ESAJ = login_data['login']['ESAJ']
        
    #     if login_ESAJ is not None:
    #         prt(type="log", message="Usuário e senha obtidos", row=1)

    #         try:
    #             self.login_esaj(driver, login_ESAJ, wait=wait)
    #         except:
    #             print(colored('Erro ao realizar login', 'red'))
    #             prt(type="error", message="Erro ao realizar login", row=1)

    #     else:
    #         print(colored('Usuário ou senha não encontrados', 'red'))
    #         prt(type="error", message="Usuário ou senha não encontrados", row=1)
    #     f.close()

    def append_process_on_output(self, processo_data, row):

        output_filename = '{dir_path}/pesquisa_processual_esaj_mov.xlsx'.format(dir_path=self.output_dir_path)

        wb = openpyxl.load_workbook(filename=output_filename)
        sheet = wb['Sheet1']

        sheet.append(processo_data)
        wb.save(output_filename)

        print(colored('Processo {} adicionado com sucesso'.format(processo_data[0]), 'green'))
        prt(type="log", message='Processo {} adicionado com sucesso'.format(processo_data[0]), row=row)

    def get_input_process(self, driver, wait):

        # clean output
        try:
            output_planilha_base = os.path.join(self.current_dir, 'pesquisa_processual_esaj_mov.xlsx')
            self.output_planilha_copia = '{dir_path}/pesquisa_processual_esaj_mov.xlsx'.format(dir_path=self.output_dir_path)

            shutil.copyfile(output_planilha_base, self.output_planilha_copia)

            output_filename = self.output_planilha_copia
            wrkbk_output = openpyxl.load_workbook(filename=output_filename)
            sheet_output = wrkbk_output.active
            sheet_output.delete_rows(2, sheet_output.max_row+1)
            wrkbk_output.save(output_filename)
        except Exception as e:
            print(e)
            print(colored('Erro ao limpar planilha de output', 'red'))
            prt(type='error', message='Erro ao limpar planilha de output', row=1)

        # iterate from each process on input
        try:
            #input_filename = os.path.join(os.getcwd(),'inputsearch', 'processos_esaj.xlsx')
            input_filename = self.input_file
            wrkbk_input = openpyxl.load_workbook(filename=input_filename)
            sheet_input = wrkbk_input.active

            for i in range(2, sheet_input.max_row+1):
                cell_obj = sheet_input.cell(row=i, column=1)
                if cell_obj.value is not None:
                    numero = cell_obj.value.replace(' ','')
                    termosLista =  sheet_input.cell(row=i, column=2).value.replace(', ',',').split(',')

                    processo_data = [numero, termosLista]

                    try:
                        self.get_processo(driver, processo_data, row=i, wait=wait)
                    except:
                        print(colored('Erro ao consultar processo {}'.format(numero), 'red'))
                        prt(type='error', message='Erro ao consultar processo {}'.format(numero), row=i)
                if i == sheet_input.max_row:
                    prt(type='log', message='Fim dos processos', row=i)
                    try:
                        os.startfile(self.output_planilha_copia)
                    except Exception as e:
                        prt(type='error', message=e)
        except:
            print(colored('Erro ao recuperar processos do input', 'red'))
            prt(type='error', message='Erro ao recuperar processos do input', row=1)

ESAJCrawler(
    credentials=sys.argv[1], 
    input_file=sys.argv[2], 
    output_dir_path=sys.argv[3]
)




