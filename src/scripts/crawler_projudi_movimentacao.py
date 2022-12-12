from concurrent.futures import process
from urllib.request import urlopen
from selenium import webdriver
from time import sleep
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options 
from selenium.webdriver.support.ui import WebDriverWait 
from selenium.webdriver.support import expected_conditions as EC
import os
import json
import openpyxl
from tkinter.scrolledtext import ScrolledText
import shutil
from termcolor import colored
import threading
import traceback
from printlog import print_log as prt
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
class PROJUDICrawler:

    def initialize_crawler(self, output_prompt, input_file, output_dir_path):

        chrome_options = Options()
        #chrome_options.add_argument("--headless")
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
        wait = WebDriverWait(driver, 10)

        self.output_prompt = output_prompt
        self.input_file = input_file
        self.output_dir_path = output_dir_path
        self.process_moves = []

        try:
            get_login_and_password = threading.Thread(target=self.get_login_and_password, args=(driver,wait))
            get_login_and_password.start()
            #self.get_login_and_password(driver)
        except:
            print(colored('Erro ao recuperar dados de login e senha', 'red'))
            prt(type="error", message='Erro ao recuperar dados de login e senha', row=1)

    def login(self, driver, login_credentials, wait):
        # 0638164-88.2019.8.04.0015
        driver.get("https://projudi.tjam.jus.br/projudi/usuario/logon.do?actionType=inicio")
        sleep(2)
        login = login_credentials['username']
        senha = login_credentials['password']

        username = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, '#login')))
        username.click()
        username.send_keys(login)
        # sleep(0.5)

        password = driver.find_element(By.CSS_SELECTOR, '#senha')
        password.send_keys(senha)
        # sleep(2)

        entrar = driver.find_element(By.CSS_SELECTOR, '#btEntrar')
        entrar.click()

        print(colored('Login efetuado com sucesso', 'green'))
        prt(type="log", message='Login efetuado com sucesso', row=1)

        try:
            self.get_input_process(driver, wait)
        except:
            print(colored('Erro ao recuperar processos do input', 'red'))
            prt(type="error", message="Erro ao recuperar processos do input", row=1)

    def get_processo(self, driver, processo_data, wait, row):

        n_processo = processo_data[0]
        driver.get("https://projudi.tjam.jus.br/projudi/processo/buscaProcessosQualquerInstancia.do?actionType=pesquisar")

        input_proc_1grau = driver.find_element(By.CSS_SELECTOR, '#numeroProcesso')

        input_proc_1grau.send_keys(n_processo)

        consultar = driver.find_element(By.CSS_SELECTOR, '#pesquisar')
        consultar.click()
        # sleep(60)
        marotagem = driver.find_element(By.CLASS_NAME, 'link')
        marotagem.click()

        try:
            self.get_process_moves(driver, processo_data, row=row, wait=wait)
        except:
            print(colored('Erro ao buscar movimentações do processo {}'.format(n_processo), 'red'))
            prt(type="error", message='Erro ao buscar movimentações do processo {}'.format(n_processo), row=row)
        
    def get_process_informations(self, driver, n_processo, row, wait):

        #sleep(10)
        print(colored('Obtendo informações do processo {}...'.format(n_processo), 'yellow'))
        prt(type="log", message='Obtendo informações do processo {}...'.format(n_processo), row=row) 

        navegar = driver.find_element(By.CSS_SELECTOR, '#tabItemprefix2')
        navegar.click()

        form = driver.find_element(By.ID, "includeContent")

        tablePoloAtivo = form.find_elements(By.CLASS_NAME, "resultTable")[0]
        nomePoloAtivo = tablePoloAtivo.find_element(By.XPATH, ".//tbody").find_elements(By.XPATH, ".//tr")[0].find_elements(By.XPATH, ".//td")[1].text.replace('(citação online)', '')
        cpfPoloAtivo = tablePoloAtivo.find_element(By.XPATH, ".//tbody").find_elements(By.XPATH, ".//tr")[0].find_elements(By.XPATH, ".//td")[3].text
        advPoloAtivo = tablePoloAtivo.find_element(By.XPATH, ".//tbody").find_elements(By.XPATH, ".//tr")[0].find_elements(By.XPATH, ".//td")[5].text

        tablePoloPassivo = form.find_elements(By.CLASS_NAME, "resultTable")[1]
        nomePoloPassivo = tablePoloPassivo.find_element(By.XPATH, ".//tbody").find_elements(By.XPATH, ".//tr")[0].find_elements(By.XPATH, ".//td")[1].text.replace('(citação online)', '')
        cpfPoloPassivo = tablePoloPassivo.find_element(By.XPATH, ".//tbody").find_elements(By.XPATH, ".//tr")[0].find_elements(By.XPATH, ".//td")[3].text

        navegar2 = driver.find_element(By.CSS_SELECTOR, '#tabItemprefix0')
        navegar2.click()

        form = driver.find_element(By.ID, "includeContent").find_element(By.CLASS_NAME, "form")

        competencia = form.find_elements(By.XPATH, ".//tr")[0].find_elements(By.XPATH, ".//td")[4].text
        vara = form.find_elements(By.XPATH, ".//tr")[1].find_elements(By.XPATH, ".//td")[4].text
        data_distribuição = form.find_elements(By.XPATH, ".//tr")[2].find_elements(By.XPATH, ".//td")[1].text
        valorDaCausa = form.find_elements(By.XPATH, ".//tr")[8].find_elements(By.XPATH, ".//td")[1].text

        processo_data = [n_processo, competencia, vara, data_distribuição, nomePoloAtivo, cpfPoloAtivo, advPoloAtivo, nomePoloPassivo, cpfPoloPassivo, valorDaCausa]

        try:
            self.append_process_on_output(processo_data, row=row)
        except:
            print(colored('Erro ao adicionar processo {} na planilha'.format(n_processo), 'red'))
            prt(type="error", message="Erro ao adicionar processo {} na planilha".format(n_processo), row=row )

    def get_process_moves(self, driver, processo_data, wait, row):
        
        n_processo = processo_data[0]

        page_size = driver.find_element(By.CSS_SELECTOR, 'select[name="pagerConfigPageSize"]')
        page_size.click()
        select_1000_items = page_size.find_element(By.CSS_SELECTOR, 'option[value="1000"]')
        select_1000_items.click()
        sleep(5)

        table_moves = driver.find_element(By.CLASS_NAME, 'resultTable')
        all_moves_from_this_page = table_moves.find_elements(By.XPATH, './/tr[contains(@class, "odd") or contains(@class, "even")][not(@style="display:none;")]')

        try:
            self.append_list_moves(driver=driver, processo_data=processo_data,list_moves=all_moves_from_this_page, row=row)
        except:
            prt(type="error", message='O processo não possui movimentações', row=row)

        try:
            self.get_process_informations(driver, n_processo, row=row, wait=wait)
        except:
            print(colored('Erro ao buscar informações do processo {}'.format(n_processo), 'red'))
            prt(type="error", message='Erro ao buscar informações do processo {}'.format(n_processo), row=row)

    def append_list_moves(self, driver, processo_data, list_moves, row):

        for move in list_moves:
            try:
                move_date = move.find_elements(By.TAG_NAME, 'td')[2].text
                move_description = move.find_elements(By.TAG_NAME, 'td')[3].text
            except:
                move_date = ''
                move_description = ''
            
            for termo in processo_data[1]:
                if termo.lower() in move_description.lower():
                    try:
                        self.process_moves.append('{move_date} - {move_description}'.format(move_date=move_date, move_description=move_description))
                    except:
                        prt(type="error", message='Erro ao append', row=row)

    def get_login_and_password(self, driver, wait):

        login_file = os.path.join(os.getcwd(),'inputsearch', 'login_dados.json')

        f = open(login_file)
        login_data = json.load(f)
        login_credentials = login_data['login']['PROJUDI']
        
        if login_credentials is not None:
            print(colored('Usuário e senha obtidos', 'green'))
            prt(type="log", message='Usuário e senha obtidos', row=1)

            try:
                self.login(driver, login_credentials, wait=wait)
            except:
                print(colored('Erro ao realizar login', 'red'))
                prt(type="error", message="Erro ao realizar login", row=1)

        else:
            print(colored('Usuário ou senha não encontrados', 'red'))
            prt(type="error", message="Usuário ou senha não encontrados", row=1)
        f.close()

    def append_process_on_output(self, processo_data, row):

        output_filename = '{dir_path}/pesquisa_processual_projudi.xlsx'.format(dir_path=self.output_dir_path)

        wb = openpyxl.load_workbook(filename=output_filename)
        sheet = wb['Sheet1']

        all_moves_string = '\n'.join(self.process_moves)
        processo_data.append(all_moves_string)
        sheet.append(processo_data)
        wb.save(output_filename)

        print(colored('Processo {} adicionado com sucesso'.format(processo_data[0]), 'green'))
        prt(type="log", message='Processo {} adicionado com sucesso'.format(processo_data[0]), row=row)

    def get_input_process(self, driver, wait):

        # clean output
        try:
            output_planilha_base = os.path.join(os.getcwd(),'outputbot', 'pesquisa_processual_projudi.xlsx')
            output_planilha_copia = '{dir_path}/pesquisa_processual_projudi.xlsx'.format(dir_path=self.output_dir_path)

            shutil.copyfile(output_planilha_base, output_planilha_copia)

            output_filename = output_planilha_copia
            wrkbk_output = openpyxl.load_workbook(filename=output_filename)
            sheet_output = wrkbk_output.active
            sheet_output.delete_rows(2, sheet_output.max_row+1)
            wrkbk_output.save(output_filename)
        except:
            print(colored('Erro ao limpar planilha de output', 'red'))
            prt(type="error", message="Erro ao limpar planilha de output", row=1)

        # iterate from each process on input
        try:
            #input_filename = os.path.join(os.getcwd(),'inputsearch', 'processos_esaj.xlsx')
            input_filename = self.input_file
            wrkbk_input = openpyxl.load_workbook(filename=input_filename)
            sheet_input = wrkbk_input.active

            for i in range(2, sheet_input.max_row+1):
                cell_obj = sheet_input.cell(row=i, column=1)
                if cell_obj.value is not None and cell_obj.value != '':
                    numero = cell_obj.value.replace(' ','')
                    termosLista =  sheet_input.cell(row=i, column=2).value.replace(', ',',').split(',')

                    processo_data = [numero, termosLista]
                    # clan moves to each process
                    self.process_moves = []

                    try:
                        self.get_processo(driver, processo_data, wait=wait, row=i)
                    except:
                        print(colored('Erro ao consultar processo {}'.format(numero), 'red'))
                        prt(type="error", message="Erro ao consultar processo {}".format(numero), row=i)
                        print(traceback.format_exc())

                if i == sheet_input.max_row:
                    self.myprint('Fim dos processos', 'green')
        except:
            print(colored('Erro ao recuperar processos do input', 'red'))
            self.myprint('Erro ao recuperar processos do input', 'red')



#start = ESAJCrawler()
#start.initialize_crawler()




