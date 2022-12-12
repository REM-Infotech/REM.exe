from selenium.webdriver import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from time import sleep
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options  
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import os
import sys
import openpyxl
import threading
from printlog import print_log as prt




class ELAWCrawler:

    def __init__(self, credentials, input_file, input_dir_path):
        chrome_options = Options()
        #chrome_options.add_argument("--headless") 

        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
        
        login, password = credentials.split('>!>')
        self.credentials = {
            'login': login,
            'password': password
        }
        self.input_file = input_file
        self.input_dir_path = input_dir_path

        prt(type='log', message='input_file', row=0)
        prt(type='log', message=input_file, row=0)

        prt(type='log', message='input_dir_path', row=0)
        prt(type='log', message=input_dir_path, row=0)

        wait = WebDriverWait(driver, 10)

        try:
            login_elaw = threading.Thread(target=self.login_elaw, args=(driver, wait))
            login_elaw.start()
            #self.login_elaw(driver)
        except:
            prt(type='error', message='Erro ao fazer login', row=0)

    def login_elaw(self, driver, wait):

        driver.get("https://amazonas.elaw.com.br/login.elaw")

        prt(type='log', message='Fazendo login no sistema...', row=0)

        # wait until page load
        username = wait.until(
            EC.presence_of_element_located((By.ID,'username')))
        username.click()
        # login = 'paula.melo4'
        login = self.credentials.get('login')
        username.send_keys(login)


        password = driver.find_element(By.CSS_SELECTOR, '#password')
        password.click()
        # senha = 'SistemaAME4@'
        senha = self.credentials.get('password')
        password.send_keys(senha)

        entrar = driver.find_element(By.ID, "j_id_a_1_5_f")
        entrar.click()
        prt(type='log', message='Login efetuado com sucesso', row=0)

        return

        try:
            self.get_processo_from_input(driver=driver, wait=wait)
        except:
            prt(type='error', message='Erro ao acessar processos do input', row=0)

    def search_processo(self, driver, andamento_data, wait, row):

        driver.get("https://amazonas.elaw.com.br/processoList.elaw")

        prt(type='log', message='Buscando processo {}...'.format(andamento_data[0]), row=row)

        numproc = wait.until(EC.presence_of_element_located((By.ID,'tabSearchTab:txtSearch')))
        numproc.click()
        numproc.clear()
        numproc.send_keys(andamento_data[0])

        search_button = driver.find_element(By.ID, 'btnPesquisar')
        search_button.click()
        search_result = wait.until(EC.presence_of_element_located((By.ID,'dtProcessoResults_data')))
        open_proc = search_result.find_element(By.ID, 'dtProcessoResults:0:btnProcesso')
        
        
        try:
            open_proc.click()
            self.add_andamento(driver=driver, andamento_data=andamento_data, wait=wait, row=row)
            prt(type='log',message='Processo encontrado',row=row)
            
        except:
            prt(type='error',message='Processo {} não encontrado'.format(andamento_data[0]), row=row)
            prt(type='error', message='Erro ao adicionar andamento',row=row)
            
    
        
            
    def add_andamento(self, driver, andamento_data, wait, row):

        prt(type = 'log', message = 'Preenchendo dados do andamento...', row=row)

        add_button = wait.until(EC.presence_of_element_located((By.ID,'tabViewProcesso:j_id_hv_4_1_3_ac:j_id_hv_4_1_3_ag_1')))
        add_button.click()

        andamento_type = wait.until(EC.presence_of_element_located((By.ID,'j_id_2n:j_id_2r_2_d')))
        andamento_type.click()

        andamento_input_filter = wait.until(EC.presence_of_element_located((By.ID,'j_id_2n:j_id_2r_2_d_filter')))
        andamento_input_filter.click()
        andamento_input_filter.send_keys(andamento_data[1])
        sleep(1)
        andamento_input_filter.send_keys(Keys.ENTER)
        sleep(2)

    

        obs = wait.until(EC.presence_of_element_located((By.ID, 'j_id_2n:txtObsAndamento')))
        obs.click()
        obs.send_keys(andamento_data[3])

        ocorrencia = wait.until(EC.presence_of_element_located((By.ID,'j_id_2n:txtOcorrenciaAndamento')))
        ocorrencia.click()
        ocorrencia.send_keys(andamento_data[2])

        prt(type='log', message='Dados do andamento preenchidos com sucesso',row=row)

        try:
            self.add_anexo(driver=driver, andamento_data=andamento_data, wait=wait, row=row)
        except:
            prt(type='error',message='Erro ao adicionar anexo', row=row)
        
    def add_anexo(self, driver, andamento_data, wait, row):

        prt(type='log', message='Adicionando anexo...', row=row)
        try:
            driver.execute_script("document.querySelector('textarea[title=\"Ocorrência \"]').blur()")
        except:
            prt(type='error', message='Erro ao carregar ocorrência', row=row)

        try:
            anexo_tipo = wait.until(EC.element_to_be_clickable((By.ID, 'j_id_2n:j_id_2r_2_2w4:eFileTipoCombo')))
            anexo_tipo.click()

            anexo_input_tipo = wait.until(EC.element_to_be_clickable((By.ID,'j_id_2n:j_id_2r_2_2w4:eFileTipoCombo_filter')))
            anexo_input_tipo.click()
            anexo_input_tipo.send_keys(andamento_data[4])
            sleep(2)
            anexo_input_tipo.send_keys(Keys.ENTER)
            sleep(2)

        except:
            prt(type='error',message='Erro ao adicionar tipo do anexo',row=row)

        try:
            file_input = driver.find_element(By.CSS_SELECTOR, '#j_id_2n\:j_id_2r_2_2w4\:j_id_2r_2_2w6_2_e_2_1_input')
            #file_path = os.path.join(os.getcwd(),'notec', 'NOTA TECNICA - {}.pdf'.format(n_processo))
            file_path = '{path}/{file_name}'.format(path=self.input_dir_path, file_name=andamento_data[5])
            file_input.send_keys(file_path)
        except:
            prt(type='error',message='Erro ao anexar. Cheque se o arquivo "{file_name}" realmente existe'.format(file_name=andamento_data[5]), row=row)

        anexo_uploaded = WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.XPATH,'//tbody[@id="j_id_2n:j_id_2r_2_2w4:gedEFileDataTable_data"]/tr/td[@role="gridcell"]')))

        prt(type='log',message='Arquivo "{file_name}" anexado com sucesso'.format(file_name=andamento_data[5]), row=row)

        try:
            self.save_andamento(driver, row=row)
        except:
            prt(type='error',message='Erro ao salvar andamento', row=row)

    def save_andamento(self, driver, row):

        prt(type='log', message='Salvando andamento...',row=row)
        save_button = driver.find_element(By.ID, 'btnSalvarAndamentoProcesso')
        save_button.click()
        sleep(3)

        prt(type='log',message='Andamento salvo',row=row)
    
    def get_processo_from_input(self, driver, wait):

        prt(type='log', message='Carregando processos do input..',row=0)
        # iterate from each process on input
        try:
            #input_filename = os.path.join(os.getcwd(),'inputsearch', 'processos_elaw.xlsx')
            input_filename = self.input_file
            wrkbk_input = openpyxl.load_workbook(filename=input_filename)
            sheet_input = wrkbk_input.active
            prt(type='log',message='Processos carregados', row=0)

            for i in range(2, sheet_input.max_row+1):
                cell_obj = sheet_input.cell(row=i, column=1)
                if cell_obj.value is not None and cell_obj.value != '':
                    numero = cell_obj.value.replace(' ','')
                    tipo_andamento = sheet_input.cell(row=i, column=2).value
                    ocorrencia = sheet_input.cell(row=i, column=3).value
                    observacao = sheet_input.cell(row=i, column=4).value
                    file_type = sheet_input.cell(row=i, column=5).value
                    file_name = sheet_input.cell(row=i, column=6).value

                    andamento_data = [numero, tipo_andamento, ocorrencia, observacao, file_type, file_name]
                


                    try:
                        self.search_processo(driver=driver, andamento_data=andamento_data, wait=wait, row=i)
                    except:
                        prt(type='error',message='Erro ao buscar processo {}'.format(numero), row=i)
                if i == sheet_input.max_row:
                    prt(type='log',message='Fim dos processos',row=0)
        except Exception as e:
            print(e)

            prt(type='error', message='Erro ao carregar processos do input',row=0)


ELAWCrawler(
    credentials=sys.argv[1], 
    input_file=sys.argv[2], 
    input_dir_path=sys.argv[3]
)
